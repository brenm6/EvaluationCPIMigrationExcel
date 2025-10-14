import openpyxl
from openpyxl.styles import Font, PatternFill

class ColumnsManager:
    def __init__(self):
        """Initialize the ExcelHeaderManager with a filename."""
    #    self.filename = filename
    #    self.workbook = openpyxl.Workbook()
    #    self.sheet = self.workbook.active
    #    self.headers = []    
        

    def set_headers(self, headers, worksheet):
        """Set the headers for the first row of the Excel sheet."""
          
        # headers_newTemplate
        fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # RGB: (255, 255, 153)   
        
        for col in range(len(headers)):
            cell = worksheet.cell(row=2, column=col + 1)  # Write to row 2
            cell.value = headers[col]
            cell.fill = fill
            cell.font = Font(bold=True)

    
        # Set sum formulas in row 1 for specific columns
        sum_columns = [
            "Anzahl von Schnittstellen                      FTP",
            "Anzahl von Schnittstellen                      SFTP"
        ]
        for col in range(len(headers)):
            header = headers[col]
            if header in sum_columns:
                # Excel columns are 1-based, so col+1
                col_letter = openpyxl.utils.get_column_letter(col + 1)
                # Sum from row 3 to the last row (assuming data starts at row 3)
                sum_formula = f"=SUM({col_letter}3:{col_letter}1048576)"
                worksheet.cell(row=1, column=col + 1).value = sum_formula
                worksheet.cell(row=1, column=col + 1).font = Font(bold=True)
            else:
                worksheet.cell(row=1, column=col + 1).value = ""        
          
# Example usage
if __name__ == "__main__":
    # Create an instance of ExcelHeaderManager
    header_manager = ColumnsManager()
    
    # Set the headers
    #header_manager.set_headers(headers_newTemplate)

    # Save the workbook
    #header_manager.save()

    # Retrieve and print the headers
    # print("Headers in the workbook:", header_manager.get_headers())