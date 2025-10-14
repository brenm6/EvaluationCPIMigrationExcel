import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

class ColumnsManager:
    def __init__(self):
        """Initialize the ExcelHeaderManager with a filename."""
    #    self.filename = filename
    #    self.workbook = openpyxl.Workbook()
    #    self.sheet = self.workbook.active
    #    self.headers = []    
        

    def set_headers(self, headers,  worksheet):
        """Set the headers for the first row of the Excel sheet."""

        # headers_newTemplate
        fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # RGB: (255, 255, 153)  
        header_font = Font(name="Arial", size=10, bold=True) 
   
        # Set the fill for the first cell in row 1 (A1)
        
        for col in range(len(headers)):
            cell = worksheet.cell(row=1, column=col + 1)
            cell.value = headers[col]
            cell.fill = fill
            cell.font = Font(bold=True)
            cell.font = header_font
            
        last_col = get_column_letter(len(headers))
        worksheet.auto_filter.ref = f"A2:{last_col}2"
        
    def first_line_bold(self, worksheet):
        """Make the first row bold."""
        header_font_big = Font(name="Arial", size=18, bold=True) 
        for col in range(1, worksheet.max_column + 1):   
            cell = worksheet.cell(row=1, column=col + 1)  
            cell.font = Font(bold=True)
            cell.font = header_font_big

    def set_colour_green(self, worksheet, columnnumber):    
    
        colour = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # RGB: 146, 208, 80   
        for row in range(3, worksheet.max_row + 1):            
            cell = worksheet.cell(row=row, column=columnnumber)            
            cell.fill = colour         
                                
    def set_clour_orange(self, worksheet, columnnumber):    
    
        colour = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
        for row in range(3, worksheet.max_row + 1):            
            cell = worksheet.cell(row=row, column=columnnumber)            
            cell.fill = colour       
            
    def set_colour_light_blue(self, worksheet, columnnumber):    
    
        colour = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        for row in range(3, worksheet.max_row + 1):            
            cell = worksheet.cell(row=row, column=columnnumber)            
            cell.fill = colour       
                
# Example usage
if __name__ == "__main__":
    # Create an instance of ExcelHeaderManager
    header_manager = ColumnsManager()
    
    # Set the headers
    #header_manager.set_headers(headers_newTemplate)

    # Save the workbook
    #header_manager.save()

    # Retrieve and print the headers
    # print("Headers in the workbook:", header_manager.get_headers())





