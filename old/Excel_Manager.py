import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import ReadOnlyWorkbookException
from openpyxl.styles import PatternFill
import time

from Columns_Manager import ColumnsManager
from Headers import headers_newTemplate
from datetime import datetime
from openpyxl.styles import Border, Side



class ExcelManager:
    def __init__(self, filename):
        self.filename = filename
        base_dir = os.path.dirname(os.path.abspath(__file__))
        #excel_path = os.path.join(base_dir, "evaluation_run_results.xlsx")
        excel_path = os.path.join(base_dir, "evaluation_run_results_DE_SAP_PA3.xlsx")
        self.workbook = openpyxl.load_workbook(excel_path)
        self.sheet = self.workbook['Full Evaluation Results']
        self.write_mode = True
        self.columns_manager = ColumnsManager()

    def create_sheet(self, title, index):
        if not self.write_mode:
            raise ReadOnlyWorkbookException('Cannot create new sheet in a read-only workbook')
        return self.workbook.create_sheet(title=title, index=index)

    def set_columns(self, sheet123):
        self.columns_manager.set_headers(headers_newTemplate, sheet123)
        sheet123.insert_rows(1) # Insert a new row at position 2 for the headers

    def fill_sheet(self, sheet123):
        full_eval = self.workbook['Full Evaluation Results']

        # --- Sortiere Full Evaluation Results nach Integration Scenario (Spalte 1) ---
        data_rows = []
        for row in range(2, full_eval.max_row + 1):
            row_data = [full_eval.cell(row=row, column=col).value for col in range(1, full_eval.max_column + 1)]
            data_rows.append(row_data)
        # Sortiere nach erstem Wert (Integration Scenario), None wird als leerer String behandelt
        data_rows.sort(key=lambda x: (str(x[0]).strip() if x[0] else ""))

        # Schreibe sortierte Daten zurück (optional: du kannst auch nur mit data_rows weiterarbeiten)
        for idx, row_data in enumerate(data_rows, start=2):
            for col, value in enumerate(row_data, start=1):
                full_eval.cell(row=idx, column=col, value=value)

        # Tabellen für die jeweiligen Spalten
        table_type_s = {}
        table_module = {}      # Nur SenderAdapterModulePresence
        table_module_r = {}    # Nur ReceiverAdapterModulePresence
        table_type_r = {}
        table_qus = {}
        table_mapping = {}
        table_udf = {}
        table_receivers_count = {}
        table_ftp_count = {}
        table_tshirt_size = {}  # Neu: Tabelle für TShirt Size


        # Aufbau der Tabellen
        for row in range(1, full_eval.max_row + 1):
            scenario = full_eval.cell(row=row, column=1).value
            rule = full_eval.cell(row=row, column=2).value
            value = full_eval.cell(row=row, column=4).value
            if scenario:
                scenario = scenario.strip()
                # Typ S
                if rule == "SenderAdapterType":
                    table_type_s.setdefault(scenario, set()).add(value)
                    # Für FTP-Count
                    if value == "FTP":
                        table_ftp_count.setdefault(scenario, 0)
                        table_ftp_count[scenario] += 1
                # Modul (nur SenderAdapterModulePresence)
                if rule == "SenderAdapterModulePresence":
                    table_module.setdefault(scenario, set()).add(rule)
                # Modul R (nur ReceiverAdapterModulePresence)
                if rule == "ReceiverAdapterModulePresence":
                    table_module_r.setdefault(scenario, set()).add(rule)
                # Typ R
                if rule == "ReceiverAdapterType":
                    table_type_r.setdefault(scenario, set()).add(value)
                    # Für FTP-Count
                    if value == "FTP":
                        table_ftp_count.setdefault(scenario, 0)
                        table_ftp_count[scenario] += 1
                # AsynchronSynchron & Quality of Service
                if rule == "SenderAdapterQoS":
                    table_qus.setdefault(scenario, set()).add(value)
                # Mapping
                if rule == "MappingType":
                    table_mapping[scenario] = value
                # UDF
                if rule == "GMMCustomUDFUsageCount":
                    table_udf[scenario] = value
                # AnzahlEmpfänger
                if rule == "ICOReceivers":
                    table_receivers_count[scenario] = value

        # Table type
        eval_by_scenario = self.workbook['Eval by Integration Scenario']
        recommendations_sheet = excel_manager.workbook['Recommendations']
        table_type = build_lookup_table(eval_by_scenario, 2, 1)
        table_tshirt_size = build_lookup_table(eval_by_scenario, 2, 4)
        table_30days = build_lookup_table(eval_by_scenario, 2, 5)
        table_min_effort = build_lookup_table(eval_by_scenario, 2, 10)
        table_max_effort = build_lookup_table(eval_by_scenario, 2, 11)
        table_avg_effort = build_lookup_table(eval_by_scenario, 2, 12)
        table_mod_category = build_lookup_table(recommendations_sheet, 2, 3)
        table_mod_item = build_lookup_table(recommendations_sheet, 2, 4)
        table_mod_recommendation = build_lookup_table(recommendations_sheet, 2, 5)  
        
        
        # --- Build mapping tables for Tasks 1-5 ---
        table_xsltx = {}
        table_javax = {}
        table_udfx = {}
        table_eoiox = {}
        table_mmx = {}

        tablvle_sftp_count = {}
        tablvle_ftps_count = {}
        table_udf_count = {}

        for row in range(2, full_eval.max_row + 1):
            scenario = full_eval.cell(row=row, column=1).value
            rule = full_eval.cell(row=row, column=2).value
            value = full_eval.cell(row=row, column=4).value
            if not scenario:
                continue
            scenario = scenario.strip()
            # SFTP/FTPS count
            if rule in ("SenderAdapterType", "ReceiverAdapterType"):
                if value == "SFTP":
                    tablvle_sftp_count.setdefault(scenario, 0)
                    tablvle_sftp_count[scenario] += 1
                if value == "FTPS":
                    tablvle_ftps_count.setdefault(scenario, 0)
                    tablvle_ftps_count[scenario] += 1
            # UDF count
            if rule == "GMMCustomUDFUsageCount":
                table_udf_count.setdefault(scenario, 0)
                if value:
                    table_udf_count[scenario] += 1

            # Task 1: XSLT
            if rule == "MappingType":
                if scenario not in table_xsltx:
                    table_xsltx[scenario] = False
                if value and "XSL" in str(value):
                    table_xsltx[scenario] = True
            # Task 2: Java
            if rule == "MappingType":
                if scenario not in table_javax:
                    table_javax[scenario] = False
                if value and "Java" in str(value):
                    table_javax[scenario] = True
            # Task 4: UDFX
            if rule == "GMMCustomUDFUsageCount":
                if scenario not in table_udfx:
                    table_udfx[scenario] = False
                if value and "GMM" in str(value):
                    table_udfx[scenario] = True
            # Task 5: EOIOX
            if rule == "SenderAdapterQoS":
                if scenario not in table_eoiox:
                    table_eoiox[scenario] = False
                if value and "GMM" in str(value):
                    table_eoiox[scenario] = True
            # Task 5: MM
            if rule == "MappingType":
                if scenario not in table_mmx:
                    table_mmx[scenario] = False
                if value and "GMM" in str(value):
                    table_mmx[scenario] = True        

        # --- Ziel-Sheet füllen ---
        last_scenario = None
        row_eval = 3
        nummer = 1
        
        
        # Create one Side instance and use it for all borders
        side = Side(style='thin')
        border = Border(left=side, right=side, top=side, bottom=side)

        
        for row in range(2, full_eval.max_row + 1):
            integration_scenario = full_eval.cell(row=row, column=1).value
            if not integration_scenario:
                continue
            integration_scenario = integration_scenario.strip()
            if integration_scenario == last_scenario:
                continue
            last_scenario = integration_scenario

            parts = integration_scenario.split("|")
            
            #Schreibe in Spalten

            add= 1;    
            # Spalte 1: Nummer
            sheet123.cell(row=row_eval, column=1, value=nummer)
            sheet123.cell(row=row_eval, column=1).border = border
            nummer += 1

            # Spalte 2: "Szenario" (Integration Scenario, Primary Key)
            sheet123.cell(row=row_eval, column=2, value=integration_scenario)
            
            # Spalte 3: "Type" (Integration Scenario, Primary Key)
            type = table_type.get(integration_scenario, "")
            if not type:
                type = "n/a"
            sheet123.cell(row=row_eval, column=2+add, value=type) 

            # Spalte 4: Message Throughput (30 Days)
            throughput = table_30days.get(integration_scenario, "")
            if not throughput:
                throughput = "n/a"
            sheet123.cell(row=row_eval, column=3+add, value=throughput)

            # Spalte 5: TShirt Size
            tshirt_size = table_tshirt_size.get(integration_scenario, "")
            if not tshirt_size:
                tshirt_size = "n/a"
            sheet123.cell(row=row_eval, column=4+add, value=tshirt_size)
            
            # Spalte 6: 
            #KI25577 #Party
            party = integration_scenario.split("|")[0] if "|" in integration_scenario else integration_scenario
            sheet123.cell(row=row_eval, column=5+add, value=party)

            # Spalte 7: System (zwischen erstem und zweitem "|")
            sender_component = parts[1] if len(parts) > 1 else ""
            sheet123.cell(row=row_eval, column=6+add, value=sender_component)        

            # Spalte 8: Sender Interface / Product CMDB (zwischen zweitem und drittem "|")
            sender_interface = parts[2] if len(parts) > 2 else ""
            sheet123.cell(row=row_eval, column=7+add, value=sender_interface)

            # Spalte 9: Typ S
            type_s_values = sorted([v for v in table_type_s.get(integration_scenario, set()) if v])
            type_s_str = " / ".join(type_s_values)
            sheet123.cell(row=row_eval, column=8+add, value=type_s_str)

            # Spalte 10: Modul (nur SenderAdapterModulePresence)
            module_present = "ja" if integration_scenario in table_module else "nein"
            sheet123.cell(row=row_eval, column=9+add, value=module_present)

            # Spalte 11: Typ R
            type_r_values = sorted([v for v in table_type_r.get(integration_scenario, set()) if v])
            type_r_str = " / ".join(type_r_values)
            sheet123.cell(row=row_eval, column=10+add, value=type_r_str)

            # Spalte 12: Modul (ReceiverAdapterModulePresence)
            module_r_present = "ja" if integration_scenario in table_module_r else "nein"
            sheet123.cell(row=row_eval, column=11+add, value=module_r_present)
            
             # Spalte 13: Receiver Component (n/a)
            sheet123.cell(row=row_eval, column=12+add, value="n/a")

            # Spalte 14: Receiver Interface (n/a)
            sheet123.cell(row=row_eval, column=13+add, value="n/a")

            # Spalte 15: AsynchronSynchron (SenderAdapterQoS, "be"->"S", "eo"->"A")
            qus_values = table_qus.get(integration_scenario, set())
            qus_str = ""
            if qus_values:
                qus_str = " / ".join(["A" if v == "eo" else "S" if v == "be" else v for v in qus_values if v])
            sheet123.cell(row=row_eval, column=14+add, value=qus_str)

            # Spalte 16: ICO (Sender Interface)
            sheet123.cell(row=row_eval, column=15+add, value=sender_interface)

            # Spalte 17: Mapping
            mapping_value = table_mapping.get(integration_scenario, "")
            sheet123.cell(row=row_eval, column=16+add, value=mapping_value)

            # Spalte 18: UDF
            udf_value = table_udf.get(integration_scenario, None)
            if udf_value == "1":
                udf_value = "ja"
            elif udf_value is None:
                udf_value = "nein"
            sheet123.cell(row=row_eval, column=17+add, value=udf_value)

            # Spalte 19: AnzahlEmpfänger
            receivers_count = table_receivers_count.get(integration_scenario, "")
            sheet123.cell(row=row_eval, column=18+add, value=receivers_count)

            # Spalte 20: Quality of Service (SenderAdapterQoS, Originalwert(e))
            qus_orig = " / ".join([v for v in qus_values if v]) if qus_values else ""
            sheet123.cell(row=row_eval, column=19+add, value=qus_orig)

            # Spalte 21: Anzahl von Schnittstellen FTP
            ftp_count = table_ftp_count.get(integration_scenario, 0)
            sheet123.cell(row=row_eval, column=20+add, value=ftp_count)

            # Spalte 22: Anzahl von Schnittstellen SFTP
            sftp_count = tablvle_sftp_count.get(integration_scenario, 0)
            sheet123.cell(row=row_eval, column=21+add, value=sftp_count)

            # Spalte 23: Anzahl von Schnittstellen FTPS
            ftps_count = tablvle_ftps_count.get(integration_scenario, 0)
            sheet123.cell(row=row_eval, column=22+add, value=ftps_count)
            
            # Spalte 24: Anzahl von Schnittstellen UDF
            udf_value_count = table_udf_count.get(integration_scenario, 0)
            sheet123.cell(row=row_eval, column=23+add, value=udf_value_count)
            
            # Spalte 25: Anzahl von Schnittstellen ABAP n/a
            # Hier wird "n/a" eingetragen, da es keine ABAP-Schnittstellen
            sheet123.cell(row=row_eval, column=24+add, value="n/a")
            
            # Spalte 26: Wahrheiswert MM vorhanden
            mmx_value = "X" if table_mmx.get(integration_scenario, False) else ""
            sheet123.cell(row=row_eval, column=25+add, value=mmx_value)

            # Spalte 27: XSLT
            xsltx_value = "X" if table_xsltx.get(integration_scenario, False) else ""
            sheet123.cell(row=row_eval, column=26+add, value=xsltx_value)

            # Spalte 28: Java
            javax_value = "X" if table_javax.get(integration_scenario, False) else ""
            sheet123.cell(row=row_eval, column=27+add, value=javax_value)

            # Spalte 29: ABAP
            sheet123.cell(row=row_eval, column=28+add, value="n/a")

            # Spalte 30: UDF
            udfx_value = "X" if table_udfx.get(integration_scenario, False) else ""
            sheet123.cell(row=row_eval, column=29+add, value=udfx_value)

            # Spalte 31: EOIO (Spalte 30 ggf. leer lassen)
            eoiox_value = "X" if table_eoiox.get(integration_scenario, False) else ""
            sheet123.cell(row=row_eval, column=30+add, value=eoiox_value)
            
            # Spalte 32: Min Effort Required (Hours)
            min_effort = table_min_effort.get(integration_scenario, "")
            if not min_effort:
                min_effort = "n/a"
            sheet123.cell(row=row_eval, column=31+add, value=min_effort)

            # Spalte 33: Max Effort Required (Hours)
            max_effort = table_max_effort.get(integration_scenario, "")
            if not max_effort:
                max_effort = "n/a"
            sheet123.cell(row=row_eval, column=32+add, value=max_effort)

            # Spalte 34: Average Effort Required (Hours)
            avg_effort = table_avg_effort.get(integration_scenario, "")
            if not avg_effort:
                avg_effort = "n/a"
            sheet123.cell(row=row_eval, column=33+add, value=avg_effort)

            # Spalte 35-37: Modernization category, Possible modernization item, Recommendation
            table_mod_category.get(integration_scenario, "n/a")
            table_mod_item.get(integration_scenario, "n/a")
            table_mod_recommendation.get(integration_scenario, "n/a")
            
            if not mod_category:
                mod_category = "n/a"
            if not mod_item:
                mod_item = "n/a"
            if not recommendation:
                recommendation = "n/a"
                        
            sheet123.cell(row=row_eval, column=34+add, value=mod_category)      
            sheet123.cell(row=row_eval, column=35+add, value=mod_item)
            sheet123.cell(row=row_eval, column=36+add, value=recommendation)

            row_eval += 1
            
        sum_columns = [
            "Anzahl von Schnittstellen                      FTP",
            "Anzahl von Schnittstellen                      SFTP",
            "Anzahl von Schnittstellen                      FTPS",
            "Anzahl von Schnittstellen                      UDF"
        ]
        count_x_columns = [
            "MM",
            "XSLT",
            "Java"
        ]

        for idx, header in enumerate(headers_newTemplate, start=1):
            col_letter = get_column_letter(idx)
            if header in sum_columns:
                # SUM for numeric columns
                sheet123.cell(row=1, column=idx).value = f"=SUM({col_letter}3:{col_letter}1048576)"
                sheet123.cell(row=1, column=idx).font = openpyxl.styles.Font(bold=True)
            elif header in count_x_columns:
                # COUNTIF for "X" in these columns
                sheet123.cell(row=1, column=idx).value = f'=COUNTIF({col_letter}3:{col_letter}1048576,"X")'
                sheet123.cell(row=1, column=idx).font = openpyxl.styles.Font(bold=True)
            else:
                sheet123.cell(row=1, column=idx).value = ""
        
        # Färbe einzelen Zeilen ein
        self.columns_manager.set_colour_green(sheet123,1)
        self.columns_manager.set_clour_orange(sheet123,6)
        self.columns_manager.set_clour_orange(sheet123,10)
        self.columns_manager.set_colour_light_blue(sheet123,15)
        
        # Erste Zeile Große Buchstaben
        self.columns_manager.first_line_bold(sheet123)               

    def save(self):
        self.workbook.save(self.filename)
 
 # nimmt nur den ersten Eintrag, wenn es zwei Zeilen mit dem gleichen Szenario gibt
# und baut eine Lookup-Tabelle auf, die für ein Szenario den Wert aus der angegeben
# Spalte zurückgibt.
# Beispiel: build_lookup_table(sheet, 2, 1) gibt für jede Zeile
# in Spalte 2 (Integration Scenario) den Wert aus Spalte 1 (Type) zurück.
# Wenn es mehrere Zeilen mit dem gleichen Szenario gibt, wird nur der erste Wert verwendet  
def build_lookup_table(sheet, key_col, value_col):
    table = {}
    for row in range(2, sheet.max_row + 1):
        key = sheet.cell(row=row, column=key_col).value
        value = sheet.cell(row=row, column=value_col).value
        if key:
            key_str = str(key).strip()
            if key_str not in table:
                table[key_str] = value
    return table

def build_recommendations_lookup(sheet):
    # Assumes: Integration Scenario = column 2, Modernization category = col 3, Possible modernization item = col 4, Recommendation = col 5
    table = {}
    for row in range(2, sheet.max_row + 1):
        key = sheet.cell(row=row, column=2).value  # Integration Scenario
        cat = sheet.cell(row=row, column=3).value  # Modernization category
        item = sheet.cell(row=row, column=4).value  # Possible modernization item
        rec = sheet.cell(row=row, column=5).value  # Recommendation
        if key:
            key_str = str(key).strip()
            table[key_str] = (
                cat if cat else "n/a",
                item if item else "n/a",
                rec if rec else "n/a"
            )
    return table


# Main Methode
if __name__ == "__main__":
    start_time = time.time()  # Start timer
    
    excel_manager = ExcelManager('interface_evaluation_result.xlsx')
    sheet_to_add = excel_manager.create_sheet('Evaluation', 2)
    excel_manager.set_columns(sheet_to_add)
    excel_manager.fill_sheet(sheet_to_add)

    # Entferne alle anderen Sheets außer "Evaluation"
    for ws in list(excel_manager.workbook.sheetnames):
        if ws != "Evaluation":
            std = excel_manager.workbook[ws]
            excel_manager.workbook.remove(std)

    excel_manager.save()
    
    end_time = time.time()  # End timer
    duration = end_time - start_time

    current_time = str(datetime.now().time().strftime("%H:%M"))
    print("evaluation_run_results_DE_SAP_PA3.xlsx safed successfully!    Time: " + current_time)
    print(f"Execution time: {duration:.2f} seconds")

