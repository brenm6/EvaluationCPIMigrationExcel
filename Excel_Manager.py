import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import ReadOnlyWorkbookException
from openpyxl.styles import PatternFill
from openpyxl.worksheet.dimensions import SheetFormatProperties
import time
from collections import defaultdict

from Columns_Manager import ColumnsManager
from Headers import headers_newTemplate
from datetime import datetime
from openpyxl.styles import Border, Side

class ExcelManager:
    def __init__(self, filename):
        self.filename = filename
        base_dir = os.path.dirname(os.path.abspath(__file__))
        excel_path = os.path.join(base_dir, "evaluation_run_results_input_PA3_2025-07-18.xlsx")
        # Use read_only=True and data_only=True for faster loading
        self.workbook = openpyxl.load_workbook(excel_path, read_only=False, data_only=True)
        self.sheet = self.workbook['Full Evaluation Results']
        self.write_mode = True
        self.columns_manager = ColumnsManager()

    def create_sheet(self, title, index):
        if not self.write_mode:
            raise ReadOnlyWorkbookException('Cannot create new sheet in a read-only workbook')
        return self.workbook.create_sheet(title=title, index=index)

    def set_columns(self, sheet123):
        self.columns_manager.set_headers(headers_newTemplate, sheet123)
        sheet123.insert_rows(1) # Insert a new row at position 2 for the headers
        
    def fill_sheet(self, sheet123):
        print("Starting optimized data processing...")
        
        # Read all data at once using bulk operations for maximum speed
        full_eval = self.workbook['Full Evaluation Results']
        eval_by_scenario = self.workbook['Eval by Integration Scenario']
        recommendations = self.workbook['Recommendations']
        
        print("Converting sheets to arrays...")
        # Convert entire sheets to arrays for faster access
        full_eval_data = []
        for row in full_eval.iter_rows(values_only=True):
            full_eval_data.append(row)
        
        eval_by_scenario_data = []
        for row in eval_by_scenario.iter_rows(values_only=True):
            eval_by_scenario_data.append(row)
            
        recommendations_data = []
        for row in recommendations.iter_rows(values_only=True):
            recommendations_data.append(row)
        
        print("Building lookup tables...")
        
        # Build lookup tables efficiently
        lookup_tables = {}
        
        # Eval by scenario lookup tables
        for i, row in enumerate(eval_by_scenario_data[1:], start=2):  # Skip header
            if row and len(row) > 1 and row[1]:  # Check if row exists and has scenario
                key = str(row[1]).strip()
                lookup_tables[key] = {
                    'type': row[0] if len(row) > 0 else None,
                    'tshirt_size': row[3] if len(row) > 3 else None,
                    '30days': row[4] if len(row) > 4 else None,
                    'min_effort': row[9] if len(row) > 9 else None,
                    'max_effort': row[10] if len(row) > 10 else None,
                    'avg_effort': row[11] if len(row) > 11 else None,
                }
        
        # Recommendations lookup tables
        recommendation_tables = {}
        for i, row in enumerate(recommendations_data[1:], start=2):  # Skip header
            if row and len(row) > 0 and row[0]:  # Check if row exists and has key
                key = str(row[0]).strip()
                recommendation_tables[key] = {
                    'mod_category': row[1] if len(row) > 1 else None,
                    'mod_item': row[2] if len(row) > 2 else None,
                    'recommendation': row[3] if len(row) > 3 else None,
                }
        
        print("Processing main evaluation data...")
        
        # Initialize all data structures
        tables = {
            'type_s': defaultdict(set),
            'module': defaultdict(set),
            'module_r': defaultdict(set), 
            'type_r': defaultdict(set),
            'scenario_values': defaultdict(list),
            'qus': defaultdict(set),
            'mapping': {},
            'udf': {},
            'functlib': {},
            'receivers_count': {},
            'ftp_count': defaultdict(int),
            'sftp_count': defaultdict(int),
            'ftps_count': defaultdict(int),
            'special_udf': {},
            'special_functlib': {},
            'rules': defaultdict(list),
            'xsltx': defaultdict(bool),
            'javax': defaultdict(bool),
            'eoiox': defaultdict(bool),
            'mmx': defaultdict(int),
        }
        
        # Special rules sets
        special_udf_rules = {
            "GMMCustomUDFDynamicConfiguration",
            "GMMCustomUDFLookupService", 
            "GMMCustomUDFFIleOS"
        }
        special_functlib_rules = {
            "GMMCustomFuncLibDynamicConfiguration",
            "GMMCustomFuncLibLookupService",
            "GMMCustomFuncLibFileOS"
        }
        
        ICOReceiversFound = False
        count = 0
        
        # Process all data in a single pass
        for row_data in full_eval_data[1:]:  # Skip header
            if not row_data or len(row_data) < 4:
                continue
                
            scenario = row_data[0]
            rule = row_data[1] 
            value = row_data[3]
            
            if not scenario:
                continue
                
            scenario = str(scenario).strip()
            
            # Add to scenario values and rules
            tables['scenario_values'][scenario].append(value)
            if rule:
                tables['rules'][scenario].append(rule)
            
            # Process rules efficiently
            if rule == "SenderAdapterType":
                tables['type_s'][scenario].add(value)
                if value == "FTP":
                    tables['ftp_count'][scenario] += 1
                elif value == "SFTP":
                    tables['sftp_count'][scenario] += 1
                elif value == "FTPS":
                    tables['ftps_count'][scenario] += 1
                    
            elif rule == "SenderAdapterModulePresence":
                tables['module'][scenario].add(rule)
                
            elif rule == "ReceiverAdapterModulePresence":
                tables['module_r'][scenario].add(rule)
                
            elif rule in ("ReceiverAdapterType", "ReceiverCustomAdapterType"):
                tables['type_r'][scenario].add(value)
                if value == "FTP":
                    tables['ftp_count'][scenario] += 1
                elif value == "SFTP":
                    tables['sftp_count'][scenario] += 1
                elif value == "FTPS":
                    tables['ftps_count'][scenario] += 1
                    
            elif rule == "SenderAdapterQoS":
                tables['qus'][scenario].add(value)
                if value and "GMM" in str(value):
                    tables['eoiox'][scenario] = True
                    
            elif rule == "MappingType":
                tables['mapping'][scenario] = value
                if value and "XSL" in str(value):
                    tables['xsltx'][scenario] = True
                if value and "Java" in str(value):
                    tables['javax'][scenario] = True
                if value and "GMM" in str(value):
                    tables['mmx'][scenario] += 1
                    
            elif rule == "GMMCustomUDFUsageCount":
                tables['udf'][scenario] = value
                
            elif rule == "GMMCustomFuncLibUsageCount":
                tables['functlib'][scenario] = value
                
            elif rule == "ICOReceivers":
                if value is not None:
                    tables['receivers_count'][scenario] = value
                    ICOReceiversFound = True
                    
            # Special rules
            if rule in special_udf_rules:
                tables['special_udf'][scenario] = True
            if rule in special_functlib_rules:
                tables['special_functlib'][scenario] = True
            
            # Handle receiver counting
            if rule in ("ReceiverAdapterType", "ReceiverCustomAdapterType") and not ICOReceiversFound:
                if value is not None:
                    count += 1
                tables['receivers_count'][scenario] = count
        
        print("Writing to output sheet...")
        
        # Get unique scenarios and sort them
        unique_scenarios = sorted(set(str(row[0]).strip() for row in full_eval_data[1:] if row and row[0]))
        
        # Pre-create styles
        side = Side(style='thin')
        border = Border(left=side, right=side, top=side, bottom=side)
        
        # Column value mapping
        column_value_map = {
            13: "AF_Modules/MessageTransformBean",
            14: "localejbs/AF_Modules/MessageLoggerBean", 
            15: "localejbs/PGPEncryption",
            16: "SAP_XI_IDOC/IDOCFlatToXmlConvertor",
            17: "AF_Modules/DynamicConfigurationBean",
            18: "AF_Modules/MultipartHeaderBean",
            19: "AF_Modules/PayloadSwapBean",
            20: "AF_Modules/XMLAnonymizerBean",
        }
        
        # Write data efficiently
        row_eval = 3
        for nummer, integration_scenario in enumerate(unique_scenarios, 1):
            scenario_data = lookup_tables.get(integration_scenario, {})
            parts = integration_scenario.split("|")
            add = 8
            
            # Column 1: Number
            sheet123.cell(row=row_eval, column=1, value=nummer).border = border
            
            # Column 2: Scenario
            sheet123.cell(row=row_eval, column=2, value=integration_scenario)
            
            # Column 3: Type
            sheet123.cell(row=row_eval, column=3, value=scenario_data.get('type', " "))
            
            # Column 4: Message Throughput (30 Days)
            throughput = scenario_data.get('30days', "n/a") or "n/a"
            sheet123.cell(row=row_eval, column=4, value=throughput)
            
            # Column 5: TShirt Size
            tshirt_size = scenario_data.get('tshirt_size', "n/a") or "n/a"
            sheet123.cell(row=row_eval, column=5, value=tshirt_size)
            
            # Column 6: Party
            party = parts[0] if parts else integration_scenario
            sheet123.cell(row=row_eval, column=6, value=party)
            
            # Column 7: System
            sender_component = parts[1] if len(parts) > 1 else " "
            sheet123.cell(row=row_eval, column=7, value=sender_component)
            
            # Column 8: Sender Interface
            sender_interface = parts[2] if len(parts) > 2 else " "
            sheet123.cell(row=row_eval, column=8, value=sender_interface)
            
            # Column 9: Type S
            type_s_values = sorted([v for v in tables['type_s'][integration_scenario] if v])
            type_s_str = " / ".join(type_s_values) if type_s_values else " "
            sheet123.cell(row=row_eval, column=9, value=type_s_str)
            
            # Column 10: Module (SenderAdapterModulePresence)
            module_present = "ja" if integration_scenario in tables['module'] else "nein"
            sheet123.cell(row=row_eval, column=10, value=module_present)
            
            # Column 11: Type R
            type_r_values = sorted([v for v in tables['type_r'][integration_scenario] if v])
            type_r_str = " / ".join(type_r_values) if type_r_values else " "
            sheet123.cell(row=row_eval, column=11, value=type_r_str)
            
            # Column 12: Module R (ReceiverAdapterModulePresence)
            module_r_present = "ja" if integration_scenario in tables['module_r'] else "nein"
            sheet123.cell(row=row_eval, column=12, value=module_r_present)
            
            # Columns 13-20: Module mappings
            for column, value_to_check in column_value_map.items():
                cell_value = "X" if any(
                    value == value_to_check
                    for value in tables['scenario_values'][integration_scenario]
                ) else " "
                sheet123.cell(row=row_eval, column=column, value=cell_value)
            
            # Additional columns with add offset
            # Receiver Component (n/a)
            sheet123.cell(row=row_eval, column=add+13, value="n/a")
            
            # Receiver Interface (n/a)
            sheet123.cell(row=row_eval, column=add+14, value="n/a")
            
            # Mapping
            mapping_value = tables['mapping'].get(integration_scenario, " ")
            sheet123.cell(row=row_eval, column=add+15, value=mapping_value)
            
            # UDF Count and processing
            udf_count = tables['udf'].get(integration_scenario, 0)
            functlib_count = tables['functlib'].get(integration_scenario, 0)
            
            # UDF processing logic
            try:
                udf_a = str(udf_count) if udf_count else ""
                udf_b = str(functlib_count) if functlib_count else ""
                
                if udf_a.strip() and udf_b.strip():
                    udf_value_str = f"{udf_a} / {udf_b}"
                elif udf_a.strip():
                    udf_value_str = udf_a
                elif udf_b.strip():
                    udf_value_str = udf_b
                else:
                    udf_value_str = " "
            except:
                udf_value_str = " "
            
            sheet123.cell(row=row_eval, column=add+18, value=udf_value_str)
            
            # Receivers count
            receivers_count = tables['receivers_count'].get(integration_scenario, " ")
            sheet123.cell(row=row_eval, column=add+19, value=receivers_count)
            
            # Quality of Service
            qus_values = tables['qus'][integration_scenario]
            qus_orig = " / ".join([v for v in qus_values if v]) if qus_values else " "
            sheet123.cell(row=row_eval, column=add+20, value=qus_orig)
            
            # FTP/SFTP/FTPS counts
            sheet123.cell(row=row_eval, column=add+21, value=tables['ftp_count'][integration_scenario])
            sheet123.cell(row=row_eval, column=add+22, value=tables['sftp_count'][integration_scenario])
            sheet123.cell(row=row_eval, column=add+23, value=tables['ftps_count'][integration_scenario])
            
            # UDF count with special rules
            udf_value_count = tables['udf'].get(integration_scenario, 0)
            try:
                if udf_value_count is not None and float(udf_value_count) >= 1:
                    sheet123.cell(row=row_eval, column=add+24, value=udf_value_count)
                elif tables['special_udf'].get(integration_scenario, False):
                    sheet123.cell(row=row_eval, column=add+24, value=1)
                else:
                    sheet123.cell(row=row_eval, column=add+24, value=0)
            except (ValueError, TypeError):
                if tables['special_udf'].get(integration_scenario, False):
                    sheet123.cell(row=row_eval, column=add+24, value=1)
                else:
                    sheet123.cell(row=row_eval, column=add+24, value=0)
            
            # FunctLib count with special rules
            functlib_value = tables['functlib'].get(integration_scenario, 0)
            try:
                if functlib_value is not None and float(functlib_value) >= 1:
                    sheet123.cell(row=row_eval, column=add+25, value=functlib_value)
                elif tables['special_functlib'].get(integration_scenario, False):
                    sheet123.cell(row=row_eval, column=add+25, value=1)
                else:
                    sheet123.cell(row=row_eval, column=add+25, value=0)
            except (ValueError, TypeError):
                if tables['special_functlib'].get(integration_scenario, False):
                    sheet123.cell(row=row_eval, column=add+25, value=1)
                else:
                    sheet123.cell(row=row_eval, column=add+25, value=0)
            
            # Dynamic Configuration, LookupService, OS (File)
            rules_list = tables['rules'][integration_scenario]
            
            # Dynamic Conf
            if "GMMCustomFuncLibDynamicConfiguration" in rules_list:
                sheet123.cell(row=row_eval, column=add+26, value="X")
            else:
                sheet123.cell(row=row_eval, column=add+26, value=" ")
            
            # LookupService
            if "GMMCustomFuncLibLookupService" in rules_list:
                sheet123.cell(row=row_eval, column=add+27, value="X")
            else:
                sheet123.cell(row=row_eval, column=add+27, value=" ")
            
            # OS (File)
            if "GMMCustomFuncLibFileOS" in rules_list:
                sheet123.cell(row=row_eval, column=add+28, value="X")
            else:
                sheet123.cell(row=row_eval, column=add+28, value=" ")
            
            # ABAP (n/a)
            sheet123.cell(row=row_eval, column=add+29, value="n/a")
            
            # Message Mappings count
            sheet123.cell(row=row_eval, column=add+30, value=tables['mmx'][integration_scenario])
            
            # XSLT, Java, ABAP, EOIO
            sheet123.cell(row=row_eval, column=add+31, value="X" if tables['xsltx'][integration_scenario] else " ")
            sheet123.cell(row=row_eval, column=add+32, value="X" if tables['javax'][integration_scenario] else " ")
            sheet123.cell(row=row_eval, column=add+33, value="n/a")
            sheet123.cell(row=row_eval, column=add+34, value="X" if tables['eoiox'][integration_scenario] else " ")
            
            # Effort columns
            min_effort = scenario_data.get('min_effort', "n/a") or "n/a"
            max_effort = scenario_data.get('max_effort', "n/a") or "n/a"
            avg_effort = scenario_data.get('avg_effort', "n/a") or "n/a"
            sheet123.cell(row=row_eval, column=add+35, value=min_effort)
            sheet123.cell(row=row_eval, column=add+36, value=max_effort)
            sheet123.cell(row=row_eval, column=add+37, value=avg_effort)
            
            # Recommendation columns
            rec_data = recommendation_tables.get(integration_scenario, {})
            mod_category = rec_data.get('mod_category', "n/a") or "n/a"
            mod_item = rec_data.get('mod_item', "n/a") or "n/a"
            recommendation = rec_data.get('recommendation', "n/a") or "n/a"
            sheet123.cell(row=row_eval, column=add+38, value=mod_category)
            sheet123.cell(row=row_eval, column=add+39, value=mod_item)
            sheet123.cell(row=row_eval, column=add+40, value=recommendation)
            
            row_eval += 1
        
        print("Applying formatting...")
        
        # Apply borders efficiently to all filled cells
        for r in range(3, row_eval):
            for c in range(1, len(headers_newTemplate) + 1):
                sheet123.cell(row=r, column=c).border = border
        
        # Add summary formulas
        sum_columns = [
            "Anzahl von Schnittstellen                      FTP",
            "Anzahl von Schnittstellen                      SFTP", 
            "Anzahl von Schnittstellen                      FTPS",
            "Anzahl von Schnittstellen                      UDF"
        ]
        count_x_columns = [
            "MM",
            "XSLT", 
            "Java"
        ]
        
        for idx, header in enumerate(headers_newTemplate, start=1):
            col_letter = get_column_letter(idx)
            if header in sum_columns:
                sheet123.cell(row=1, column=idx).value = f"=SUM({col_letter}3:{col_letter}1048576)"
                sheet123.cell(row=1, column=idx).font = openpyxl.styles.Font(bold=True)
            elif header in count_x_columns:
                sheet123.cell(row=1, column=idx).value = f'=COUNTIF({col_letter}3:{col_letter}1048576,"X")'
                sheet123.cell(row=1, column=idx).font = openpyxl.styles.Font(bold=True)
            else:
                sheet123.cell(row=1, column=idx).value = " "
        
        # Apply column formatting
        self.columns_manager.set_column_width(sheet123, 7, 20)
        self.columns_manager.set_column_width(sheet123, 8, 20)
        self.columns_manager.set_column_width(sheet123, 25, 25)
        
        # Apply colors
        color_columns_green = [1, 13, 14, 15, 16, 17, 18, 19, 20]
        for col in color_columns_green:
            self.columns_manager.set_colour_green(sheet123, col)
        
        color_columns_orange = [6, 10]
        for col in color_columns_orange:
            self.columns_manager.set_clour_orange(sheet123, col)
        
        color_columns_blue = [35, 36, 37]
        for col in color_columns_blue:
            self.columns_manager.set_colour_light_blue(sheet123, col)
        
        # Make first line bold
        self.columns_manager.first_line_bold(sheet123)
        
        print("Data processing completed successfully.")
    
    
    def group_columns(self):
        self.sheet.column_dimensions.group("Y", "AA", hidden=True)  # Nummer

    def save(self):
        self.workbook.save("evaluation_run_results_DE_SAP_PA3.xlsx")
 
# nimmt nur den ersten Eintrag, wenn es zwei Zeilen mit dem gleichen Szenario gibt
# und baut eine Lookup-Tabelle auf, die für ein Szenario den Wert aus der angegeben
# Spalte zurückgibt.
# Beispiel: build_lookup_table(sheet, 2, 1) gibt für jede Zeile
# in Spalte 2 (Integration Scenario) den Wert aus Spalte 1 (Type) zurück.
# Wenn es mehrere Zeilen mit dem gleichen Szenario gibt, wird nur der erste Wert verwendet  
def build_lookup_table(sheet, key_col, value_col):
    table = {}
    for row in range(2, sheet.max_row + 1):
        key = sheet.cell(row=row, column=key_col).value
        value = sheet.cell(row=row, column=value_col).value

        if key:
            key_str = str(key).strip()
            if key_str not in table:
                table[key_str] = value
    return table


# Main Methode
if __name__ == "__main__":
    
    # The following block is intentionally commented out.
    # try:
    #     from Frontend import ExcelFrontend  # Import hier lokal!
    # except ImportError:
    #     raise ImportError("The module 'frontend' could not be found. Please ensure 'frontend.py' exists in the same directory as this script.")
    # import tkinter as tk
    # root = tk.Tk()
    # app = ExcelFrontend(root)
    # root.mainloop()

# die Verarbeitung der Excel-Datei OHNE Frontent 
# die Logik liegt jetzt im Frontend
    print("Starte Verarbeitung der Excel-Datei OHNE Frontend...")
    start_time = time.time()  # Start timer
    
    excel_manager = ExcelManager('interface_evaluation_result.xlsx')
    sheet_to_add = excel_manager.create_sheet('Evaluation', 2)
    
    excel_manager.set_columns(sheet_to_add)
    excel_manager.fill_sheet(sheet_to_add)
    excel_manager.group_columns()

    # Entferne alle anderen Sheets außer "Evaluation"
    for ws in list(excel_manager.workbook.sheetnames):
        if ws != "Evaluation":
            std = excel_manager.workbook[ws]
            excel_manager.workbook.remove(std)

    excel_manager.save()
    
    end_time = time.time()  # End timer
    duration = end_time - start_time

    current_time = str(datetime.now().time().strftime("%H:%M"))
    print("evaluation_run_results_DE_SAP_PA3.xlsx saved successfully!    Time: " + current_time)
    print(f"Execution time: {duration:.2f} seconds")